"""
Convert Excel data back to grantha-details.json
Usage: python excel_to_json.py
Input: grantha-details.xlsx
Output: Grantha/grantha-details-new.json (backup original)
"""

import json
import openpyxl
from datetime import datetime
import os
import shutil

# Backup original JSON
print("Creating backup of original JSON...")
original_file = 'Grantha/grantha-details.json'
if os.path.exists(original_file):
    backup_file = f'Grantha/grantha-details-backup-{datetime.now().strftime("%Y%m%d-%H%M%S")}.json'
    shutil.copy2(original_file, backup_file)
    print(f"[OK] Backup created: {backup_file}")

# Read Excel file
excel_file = "grantha-details.xlsx"
if not os.path.exists(excel_file):
    print(f"[ERROR] ERROR: {excel_file} not found!")
    print("Please make sure the Excel file exists.")
    exit(1)

print(f"Reading {excel_file}...")
wb = openpyxl.load_workbook(excel_file)

# Find data sheet (skip instructions sheet)
data_sheet = None
for sheet_name in wb.sheetnames:
    if "Vadavali Data" in sheet_name or "Data" in sheet_name:
        data_sheet = wb[sheet_name]
        break

if not data_sheet:
    # Use first sheet if no specific name found
    data_sheet = wb.active

print(f"Processing sheet: {data_sheet.title}")

# Read headers
headers = []
for col_idx in range(1, 8):  # A to G (added Topic Title column)
    header = data_sheet.cell(row=1, column=col_idx).value
    headers.append(header)

print(f"Headers: {headers}")

# Validate headers (Topic Title is optional reference column)
required_headers = ["Topic ID", "Topic Title (Reference)", "Part", "वादावली", "भावदीपा", "प्रकाशः", "विवर्णम्"]
if headers[:7] != required_headers:
    print("[WARNING]  WARNING: Headers don't match expected format!")
    print(f"Expected: {required_headers}")
    print(f"Found:    {headers[:6]}")
    response = input("Continue anyway? (y/n): ")
    if response.lower() != 'y':
        exit(1)

# Convert to JSON structure
data = {}
row_count = 0
error_count = 0

for row_idx in range(2, data_sheet.max_row + 1):
    topic_id = str(data_sheet.cell(row=row_idx, column=1).value).strip()
    # Skip column 2 (Topic Title - reference only, not used in JSON)
    part = str(data_sheet.cell(row=row_idx, column=3).value).strip()

    # Skip empty rows
    if not topic_id or topic_id == "None" or not part:
        continue

    # Validate part format
    if not part.startswith("Part#"):
        print(f"[WARNING]  Row {row_idx}: Invalid part format '{part}' (should be Part#1, Part#2, etc.)")
        error_count += 1
        continue

    # Initialize topic if not exists
    if topic_id not in data:
        data[topic_id] = {}

    # Initialize part if not exists
    if part not in data[topic_id]:
        data[topic_id][part] = {}

    # Read commentary data (columns shifted by 1 due to Topic Title reference column)
    vadavali = data_sheet.cell(row=row_idx, column=4).value or ""
    bhavadipa = data_sheet.cell(row=row_idx, column=5).value or ""
    prakasha = data_sheet.cell(row=row_idx, column=6).value or ""
    vivarnam = data_sheet.cell(row=row_idx, column=7).value or ""

    # Store in JSON structure
    if vadavali:
        data[topic_id][part]["वादावली"] = vadavali.strip()
    if bhavadipa:
        data[topic_id][part]["भावदीपा"] = bhavadipa.strip()
    if prakasha:
        data[topic_id][part]["प्रकाशः"] = prakasha.strip()
    if vivarnam:
        data[topic_id][part]["विवर्णम्"] = vivarnam.strip()

    row_count += 1

# Sort topics by numeric ID
sorted_data = {}
for topic_id in sorted(data.keys(), key=lambda x: int(x)):
    sorted_data[topic_id] = data[topic_id]

# Write to new JSON file
output_file = 'Grantha/grantha-details-new.json'
with open(output_file, 'w', encoding='utf-8') as f:
    json.dump(sorted_data, f, ensure_ascii=False, indent=2)

print(f"\n[OK] Successfully converted Excel to JSON!")
print(f" Processed {row_count} rows")
if error_count > 0:
    print(f"[WARNING]  {error_count} errors found (check warnings above)")
print(f" New file created: {output_file}")
print(f"\n[WARNING]  IMPORTANT: Review the new file before replacing the original!")
print(f"If everything looks good, rename:")
print(f"  {output_file} → {original_file}")
