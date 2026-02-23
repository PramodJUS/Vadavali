"""
Convert grantha-details.json to Excel for easy data entry
Usage: python json_to_excel.py
Output: grantha-details.xlsx
"""

import json
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# Read JSON file
print("Reading grantha-details.json...")
with open('Grantha/grantha-details.json', 'r', encoding='utf-8') as f:
    data = json.load(f)

# Read topic titles from mainpage.csv
print("Reading topic titles from mainpage.csv...")
topic_titles = {}
import csv
with open('Grantha/mainpage.csv', 'r', encoding='utf-8') as f:
    reader = csv.DictReader(f)
    for row in reader:
        topic_titles[row['id']] = row['sutra_text']

# Create Excel workbook
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Vadavali Data"

# Create instructions sheet
instructions_ws = wb.create_sheet("📖 Instructions", 0)
instructions = [
    ["Vadavali Data Entry Instructions", ""],
    ["", ""],
    ["Column Guide:", ""],
    ["1. Topic ID", "Topic number (1, 2, 3, etc.) - DO NOT MODIFY"],
    ["2. Topic Title", "Sanskrit title for reference - READ ONLY (grayed out)"],
    ["3. Part", "Part number (Part#1, Part#2) - DO NOT MODIFY"],
    ["4-7. Commentary columns", "Fill Sanskrit text in these columns"],
    ["", ""],
    ["How to fill data:", ""],
    ["• Topic Title column is REFERENCE ONLY - it shows what topic you're working on", ""],
    ["• Fill commentary text in columns 4-7 (वादावली, भावदीपा, प्रकाशः, विवर्णम्)", ""],
    ["• For line breaks, use Alt+Enter (Windows) or Option+Enter (Mac)", ""],
    ["• Copy-paste from Word/PDF is fine", ""],
    ["• Keep Topic ID and Part exactly as shown", ""],
    ["", ""],
    ["[WARNING] IMPORTANT:", ""],
    ["• Don't modify Topic ID or Part columns", ""],
    ["• Don't modify column headers", ""],
    ["• Save as .xlsx format", ""],
    ["• Topic Title is just for your reference - it won't be saved to JSON", ""],
    ["", ""],
    ["After filling:", ""],
    ["1. Save this file", ""],
    ["2. Run: python excel_to_json.py", ""],
    ["3. Your grantha-details.json will be updated", ""],
]

for row_idx, row_data in enumerate(instructions, 1):
    for col_idx, cell_value in enumerate(row_data, 1):
        cell = instructions_ws.cell(row=row_idx, column=col_idx)
        cell.value = cell_value
        if row_idx == 1:
            cell.font = Font(bold=True, size=14, color="FFFFFF")
            cell.fill = PatternFill(start_color="2563eb", end_color="2563eb", fill_type="solid")
        elif cell_value and cell_value.startswith(("How", "⚠️", "After")):
            cell.font = Font(bold=True, size=12)

instructions_ws.column_dimensions['A'].width = 25
instructions_ws.column_dimensions['B'].width = 60

# Headers for data sheet (added Topic Title as reference)
headers = ["Topic ID", "Topic Title (Reference)", "Part", "वादावली", "भावदीपा", "प्रकाशः", "विवर्णम्"]
header_colors = ["2563eb", "16a34a", "2563eb", "dc2626", "ea580c", "9333ea", "0891b2"]

for col_idx, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col_idx)
    cell.value = header
    cell.font = Font(bold=True, size=12, color="FFFFFF")
    cell.fill = PatternFill(start_color=header_colors[col_idx-1], end_color=header_colors[col_idx-1], fill_type="solid")
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

# Set column widths
ws.column_dimensions['A'].width = 10  # Topic ID
ws.column_dimensions['B'].width = 40  # Topic Title (Reference)
ws.column_dimensions['C'].width = 12  # Part
for col_idx in range(4, 8):  # Commentary columns
    ws.column_dimensions[get_column_letter(col_idx)].width = 50

# Fill data
print("Converting data to Excel format...")
row_num = 2
for topic_id in sorted(data.keys(), key=lambda x: int(x)):
    topic_data = data[topic_id]

    # Get all parts
    parts = sorted([p for p in topic_data.keys() if p.startswith("Part#")],
                   key=lambda x: int(x.replace("Part#", "")))

    for part in parts:
        part_data = topic_data[part]

        # Get topic title for reference
        topic_title = topic_titles.get(topic_id, "")

        row_data = [
            topic_id,
            topic_title,  # Reference only - will be grayed out
            part,
            part_data.get("वादावली", ""),
            part_data.get("भावदीपा", ""),
            part_data.get("प्रकाशः", ""),
            part_data.get("विवर्णम्", "")
        ]

        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_idx)
            cell.value = value
            cell.alignment = Alignment(vertical="top", wrap_text=True)

            # Gray out Topic Title column (reference only)
            if col_idx == 2:
                cell.fill = PatternFill(start_color="f3f4f6", end_color="f3f4f6", fill_type="solid")
                cell.font = Font(italic=True, color="6b7280")

        row_num += 1

# Freeze header row
ws.freeze_panes = "A2"

# Save Excel file
output_file = "grantha-details.xlsx"
wb.save(output_file)
print(f"[SUCCESS] Created {output_file}")
print(f"Total rows: {row_num - 2}")
print("\nOpen 'grantha-details.xlsx' to start data entry!")
